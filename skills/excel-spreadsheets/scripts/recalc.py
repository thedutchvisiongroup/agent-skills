"""
Recalculate all formulas in an Excel file via LibreOffice headless, then
scan for residual errors (#REF!, #DIV/0!, etc.).

Usage:
    python recalc.py <excel_file> [timeout_seconds]
"""

import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

from _soffice import _find_soffice, macro_dir, run_soffice, soffice_env
from openpyxl import load_workbook

RECALC_MACRO = """\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" \
script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

MACRO_SCRIPT_URI = "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application"

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")

TABLE_STYLE_RE = re.compile(rb"<tableStyleInfo[^/>]*/>")

MAX_ERROR_LOCATIONS = 20

DEFAULT_TIMEOUT = 30


def _ensure_macro() -> bool:
    """Install the recalculation macro into the LibreOffice user profile.

    If the macro file already exists and contains the expected subroutine,
    this is a no-op.  Otherwise it creates the macro directory (booting
    LibreOffice once if needed to initialise the user profile) and writes
    the ``RecalculateAndSave`` macro.

    :returns: ``True`` if the macro is ready, ``False`` on write failure.
    """
    macro_file = macro_dir() / "Module1.xba"

    if macro_file.exists() and "RecalculateAndSave" in macro_file.read_text(encoding="utf-8"):
        return True

    if not macro_file.parent.exists():
        import subprocess

        subprocess.run(
            [_find_soffice(), "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=soffice_env(),
        )
        macro_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        macro_file.write_text(RECALC_MACRO, encoding="utf-8")
        return True
    except Exception:
        return False


def _snapshot_table_styles(path: Path) -> dict[str, bytes]:
    """Capture ``<tableStyleInfo>`` elements before LibreOffice overwrites them.

    LibreOffice strips Excel-specific table style info when it saves.  This
    function reads the raw XML from the ``.xlsx`` zip archive so the styles
    can be restored after recalculation via :func:`_restore_table_styles`.

    :param path: Path to the ``.xlsx`` file.
    :returns: Mapping of ``xl/tables/*.xml`` zip entry names to their
              ``<tableStyleInfo …/>`` bytes.
    """
    styles: dict[str, bytes] = {}
    with zipfile.ZipFile(path, "r") as zf:
        for name in zf.namelist():
            if name.startswith("xl/tables/") and name.endswith(".xml"):
                m = TABLE_STYLE_RE.search(zf.read(name))
                if m:
                    styles[name] = m.group(0)
    return styles


def _patch_table_style(data: bytes, style_element: bytes) -> bytes:
    """Replace or insert a ``<tableStyleInfo>`` element in table XML.

    :param data: Raw XML bytes of a single table definition.
    :param style_element: The ``<tableStyleInfo …/>`` bytes to restore.
    :returns: Patched XML bytes.
    """
    if TABLE_STYLE_RE.search(data):
        return TABLE_STYLE_RE.sub(style_element, data)
    return data.replace(b"</table>", style_element + b"</table>")


def _restore_table_styles(path: Path, styles: dict[str, bytes]) -> None:
    """Restore ``<tableStyleInfo>`` elements that LibreOffice stripped.

    Rewrites the ``.xlsx`` zip archive in-place, patching only the table
    XML entries captured by :func:`_snapshot_table_styles`.

    :param path: Path to the ``.xlsx`` file (modified in-place).
    :param styles: Mapping returned by :func:`_snapshot_table_styles`.
    """
    if not styles:
        return

    tmp = path.parent / (path.name + ".tmp")
    try:
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in styles:
                    data = _patch_table_style(data, styles[item.filename])
                zout.writestr(item, data)
        shutil.move(str(tmp), str(path))
    except Exception:
        if tmp.exists():
            tmp.unlink()


def _scan_errors(path: Path) -> dict[str, list[str]]:
    """Open the recalculated workbook and scan every cell for Excel errors.

    Uses ``data_only=True`` so formula cells contain their computed values
    (or error strings) rather than the formula text.

    :param path: Path to the ``.xlsx`` file.
    :returns: Mapping of error type (e.g. ``"#REF!"``) to a list of
              ``"SheetName!CellRef"`` location strings.
    """
    wb = load_workbook(path, data_only=True)
    found: dict[str, list[str]] = {e: [] for e in EXCEL_ERRORS}
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            found[err].append(f"{sheet_name}!{cell.coordinate}")
                            break
    wb.close()
    return found


def _count_formulas(path: Path) -> int:
    """Count all formula cells in the workbook.

    Uses ``data_only=False`` so formula strings are visible.

    :param path: Path to the ``.xlsx`` file.
    :returns: Total number of cells whose value starts with ``=``.
    """
    wb = load_workbook(path, data_only=False)
    count = 0
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    count += 1
    wb.close()
    return count


def recalc(filename: str, timeout: int = DEFAULT_TIMEOUT) -> dict:
    """Recalculate all formulas in *filename* and scan for errors.

    Workflow:
    1. Ensure the LibreOffice recalculation macro is installed.
    2. Snapshot Excel table styles (LibreOffice strips them on save).
    3. Run LibreOffice headless to recalculate and save.
    4. Restore the table styles.
    5. Scan every cell for residual Excel errors.

    :param filename: Path to the ``.xlsx`` file.
    :param timeout: Maximum seconds for the LibreOffice process.
    :returns: A dict with ``status``, ``total_errors``, ``total_formulas``,
              and ``error_summary`` on success, or ``error`` on failure.
    """
    path = Path(filename).resolve()
    if not path.exists():
        return {"error": f"File {filename} does not exist"}

    if not _ensure_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    try:
        table_styles = _snapshot_table_styles(path)
    except Exception:
        table_styles = {}

    result = run_soffice(["--headless", "--norestore", MACRO_SCRIPT_URI, str(path)], timeout=timeout)

    if result.returncode not in (0, 124):
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        _restore_table_styles(path, table_styles)
    except Exception:
        pass

    try:
        errors = _scan_errors(path)
        total_errors = sum(len(locs) for locs in errors.values())

        summary = {}
        for err_type, locations in errors.items():
            if locations:
                summary[err_type] = {
                    "count": len(locations),
                    "locations": locations[:MAX_ERROR_LOCATIONS],
                }

        return {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": _count_formulas(path),
            "error_summary": summary,
        }
    except Exception as e:
        return {"error": str(e)}


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_TIMEOUT
    print(json.dumps(recalc(filename, timeout), indent=2))


if __name__ == "__main__":
    main()
